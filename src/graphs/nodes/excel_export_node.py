import os
import io
from typing import Any, Dict, List
from langchain_core.runnables import RunnableConfig
from langgraph.runtime import Runtime
from coze_coding_utils.runtime_ctx.context import Context
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from coze_coding_dev_sdk.s3 import S3SyncStorage
from graphs.state import ExcelExportInput, ExcelExportOutput


def excel_export_node(
    state: ExcelExportInput, config: RunnableConfig, runtime: Runtime[Context]
) -> ExcelExportOutput:
    """
    title: Excel文件导出
    desc: 将校验后的交易记录导出为统一格式的Excel文件，问题行标黄
    integrations: S3 Storage, openpyxl
    """
    ctx = runtime.context

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "交易明细"

    # 定义表头
    headers = [
        "序号", "开户行", "账号", "目前状态", "发生往来日期",
        "收入", "支出", "币种", "银行流水摘要", "交易对方名称",
        "交易对方账户", "交易对手方与被核查人关系", "往来原因分析",
        "核查方式（供参考，无需填写）", "取得底稿（必须填写）", "核查备注",
        "底稿索引", "核查人", "余额"
    ]

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 定义黄色填充（用于标记问题行）
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 处理银行名称和账号
    bank_name = state.bank_name if state.bank_name else ""
    account_number = state.account_number if state.account_number else ""

    # 写入数据行
    for idx, transaction in enumerate(state.validated_transactions, 1):
        row_num = idx + 1

        # 判断该行是否有问题
        is_valid = transaction.get("is_valid", True)
        validation_note = transaction.get("validation_note", "")

        # 准备行数据
        row_data = [
            idx,  # 序号
            bank_name,  # 开户行
            account_number,  # 账号
            "正常",  # 目前状态
            transaction.get("transaction_date", ""),  # 发生往来日期
            transaction.get("income", 0) if transaction.get("income") is not None else 0,  # 收入
            transaction.get("expense", 0) if transaction.get("expense") is not None else 0,  # 支出
            transaction.get("currency", "人民币"),  # 币种
            transaction.get("summary", ""),  # 银行流水摘要
            transaction.get("counterpart_name", ""),  # 交易对方名称
            transaction.get("counterpart_account", ""),  # 交易对方账户
            transaction.get("relationship", ""),  # 交易对手方与被核查人关系
            transaction.get("reason", ""),  # 往来原因分析
            "",  # 核查方式
            "",  # 取得底稿
            validation_note if validation_note else "",  # 核查备注
            "",  # 底稿索引
            "",  # 核查人
            transaction.get("balance", "") if transaction.get("balance") is not None else ""  # 余额
        ]

        # 写入单元格
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")

            # 如果该行有问题，标黄
            if not is_valid or validation_note:
                cell.fill = yellow_fill

    # 自动调整列宽
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 15

    # 调整特定列宽
    ws.column_dimensions['A'].width = 8  # 序号
    ws.column_dimensions['E'].width = 20  # 发生往来日期
    ws.column_dimensions['I'].width = 25  # 银行流水摘要
    ws.column_dimensions['J'].width = 20  # 交易对方名称
    ws.column_dimensions['K'].width = 25  # 交易对方账户
    ws.column_dimensions['P'].width = 30  # 核查备注
    ws.column_dimensions['S'].width = 15  # 余额

    # 保存到内存
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_content = excel_buffer.getvalue()

    # 上传到对象存储
    storage = S3SyncStorage(
        endpoint_url=os.getenv("COZE_BUCKET_ENDPOINT_URL"),
        access_key="",
        secret_key="",
        bucket_name=os.getenv("COZE_BUCKET_NAME"),
        region="cn-beijing",
    )

    # 生成文件名
    safe_bank_name = bank_name.replace(" ", "_").replace("/", "_") if bank_name else "Wei_Zhi_Yin_Xing"
    file_name = f"{safe_bank_name}_银行流水.xlsx"

    # 上传文件
    file_key = storage.upload_file(
        file_content=excel_content,
        file_name=file_name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # 生成签名URL
    excel_url = storage.generate_presigned_url(
        key=file_key,
        expire_time=86400  # 24小时有效
    )

    return ExcelExportOutput(
        excel_url=excel_url,
        transaction_count=len(state.validated_transactions)
    )
